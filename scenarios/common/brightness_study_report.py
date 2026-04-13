# Copyright (c) Microsoft. All rights reserved.
# Licensed under the MIT license. See LICENSE file in the project root for full license information.

##
# brightness_study_report
#
# Generates a brightness curve study report from BacklightSlider study data.
# Scans Brightness-XX folders, reads *_metrics.csv files, and creates an XLSX
# with a Backlight Curve sheet, Power Summary, and per-brightness sheets.
#
# Parameters:
#   result_path      Path to TR-BrightnessSlider folder (default: auto-detect from study_result_dir)
#   name             Output report filename (default: auto-generated from study name + device name)
#   device_name      Device name used in report filename
#   backlight_key    Metric name for backlight power (default: "DisplayLight Power (W)")
#   analog_key       Metric name for analog/panel power (default: "DisplayLogic Power (W)")
##

import os
import re
import fnmatch
import csv
import logging
import collections

import openpyxl
from openpyxl.chart import Reference, Series, ScatterChart
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

import core.app_scenario
from core.parameters import Params


def _get_brightness_from_path(path):
    """Extract brightness value from a path containing a Brightness-XX folder."""
    parts = os.path.normpath(path).split(os.sep)
    for part in parts:
        match = re.match(r'^Brightness-(\d+)$', part)
        if match:
            return int(match.group(1))
    return None


def _convert_type(s):
    try:
        return int(s)
    except Exception:
        pass
    try:
        return float(s)
    except Exception:
        return s


def _collect_metrics(dirpath):
    """Walk dirpath and collect *_metrics.csv files into an OrderedDict keyed by brightness level."""
    metrics = collections.OrderedDict()
    basepath = "*metrics.csv"

    for root, _, files in os.walk(dirpath):
        parent = os.path.abspath(os.path.join(root, ".."))
        parent_parent = os.path.abspath(os.path.join(parent, ".."))

        if not os.path.exists(os.path.join(root, ".PASS")) and \
           not os.path.exists(os.path.join(parent, ".PASS")) and \
           not os.path.exists(os.path.join(parent_parent, ".PASS")):
            continue

        for file in files:
            if fnmatch.fnmatch(file, basepath):
                if any(skip in root.lower() for skip in ["config_check", "training", "misc", "prep", "fail"]):
                    continue
                inputfile = os.path.join(root, file)

                if parent == dirpath:
                    runtype = "Power"
                else:
                    if os.path.exists(os.path.join(root, ".PASS")):
                        runtype = os.path.basename(parent)
                    else:
                        runtype = os.path.basename(os.path.dirname(parent_parent))

                run_name = file[:-12]
                brightness_val = _get_brightness_from_path(root)
                if brightness_val is not None:
                    test_name = f"brightness_{brightness_val:03d}"
                else:
                    test_name = run_name[:-4]

                compound_name = runtype + "." + test_name
                if compound_name not in metrics:
                    metrics[compound_name] = collections.OrderedDict()

                logging.info("Reading metrics file: %s", inputfile)
                if 'Metric' not in metrics[compound_name]:
                    metrics[compound_name]['Metric'] = []
                metrics[compound_name]['Metric'].append(run_name[-3:])
                num_runs = len(metrics[compound_name]['Metric'])

                with open(inputfile) as csvfile:
                    reader = csv.reader(csvfile, delimiter=',', quotechar='"')
                    for r in reader:
                        if len(r) < 2:
                            continue
                        key = r[0]
                        val = _convert_type(r[1])
                        if key not in metrics[compound_name]:
                            metrics[compound_name][key] = []
                        num_this_metric = len(metrics[compound_name][key])
                        for i in range(num_this_metric, num_runs - 1):
                            metrics[compound_name][key].append("NA")
                        metrics[compound_name][key].append(val)

    # Sort by compound name for correct brightness ordering
    sorted_metrics = collections.OrderedDict()
    for key in sorted(metrics.keys()):
        sorted_metrics[key] = metrics[key]
    return sorted_metrics


def _create_backlight_curve_sheet(wb, metrics, backlight_key, analog_key):
    """Create a Backlight Curve sheet with data table and power curve chart."""
    ws = wb.create_sheet("Backlight Curve", 0)

    col_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    col_header_font = openpyxl.styles.Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    number_fmt = "0.000"

    power_metric_map = {
        backlight_key: "Backlight Power (W)",
        analog_key: "Analog Power (W)",
    }

    brightness_pattern = re.compile(r'^brightness_(\d+)$')
    data_points = []

    for compound_name in metrics:
        if '.' not in compound_name:
            continue
        runtype, testname = compound_name.split('.', 1)
        match = brightness_pattern.match(testname)
        if not match:
            continue

        brightness_pct = int(match.group(1))
        point = {"brightness_pct": brightness_pct}

        for src_key, dst_key in power_metric_map.items():
            if src_key in metrics[compound_name]:
                vals = metrics[compound_name][src_key]
                numeric_vals = [v for v in vals if isinstance(v, (int, float))]
                if numeric_vals:
                    point[dst_key] = round(sum(numeric_vals) / len(numeric_vals), 6)
                else:
                    point[dst_key] = None
            else:
                point[dst_key] = None

        bl = point.get("Backlight Power (W)")
        an = point.get("Analog Power (W)")
        if bl is not None and an is not None:
            point["Total Display Power (W)"] = round(bl + an, 6)
        else:
            point["Total Display Power (W)"] = None

        data_points.append(point)

    data_points.sort(key=lambda x: x["brightness_pct"])

    if not data_points:
        ws.cell(row=1, column=1, value="No brightness data found.")
        return

    row = 1
    ws.cell(row=row, column=1, value="Backlight Curve Report")
    ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
    row += 2

    table_start_row = row
    columns = [
        ("Bucket", 23),
        ("Slider %", 23),
        ("Backlight Power (W)", 22),
        ("Analog Power (W)", 20),
        ("Total Display Power (W)", 24),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    row += 1
    data_start_row = row

    for i, point in enumerate(data_points):
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=point["brightness_pct"]).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

        for col_idx, key in [(3, "Backlight Power (W)"),
                              (4, "Analog Power (W)"),
                              (5, "Total Display Power (W)")]:
            cell = ws.cell(row=row, column=col_idx, value=point.get(key))
            cell.number_format = number_fmt
            cell.border = thin_border
        row += 1

    data_end_row = row - 1

    row += 1
    ws.cell(row=row, column=1, value="Summary")
    ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    by_pct = {p["brightness_pct"]: p for p in data_points}
    for pct in [0, 50, 100]:
        if pct in by_pct:
            p = by_pct[pct]
            ws.cell(row=row, column=1, value=f"At {pct}% Brightness")
            ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
            ws.cell(row=row, column=2, value="Total Display Power (W)")
            cell = ws.cell(row=row, column=3, value=p.get("Total Display Power (W)"))
            cell.number_format = number_fmt
            row += 1

    display_powers = [p["Total Display Power (W)"] for p in data_points
                      if p.get("Total Display Power (W)") is not None]
    if display_powers:
        row += 1
        ws.cell(row=row, column=1, value="Min Display Power (W)")
        ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row, column=2, value=min(display_powers))
        ws.cell(row=row, column=2).number_format = number_fmt
        row += 1
        ws.cell(row=row, column=1, value="Max Display Power (W)")
        ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row, column=2, value=max(display_powers))
        ws.cell(row=row, column=2).number_format = number_fmt
        row += 1
        ws.cell(row=row, column=1, value="Delta (W)")
        ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row, column=2, value=round(max(display_powers) - min(display_powers), 6))
        ws.cell(row=row, column=2).number_format = number_fmt

    chart = ScatterChart()
    chart.title = "Brightness Slider Power Curve"
    chart.x_axis.title = "Slider %"
    chart.y_axis.title = "Power (W)"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 100
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.majorUnit = 10
    chart.x_axis.delete = False
    chart.y_axis.scaling.min = 0
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.delete = False
    chart.y_axis.numFmt = "0.0"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    x_values = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)

    y_total = Reference(ws, min_col=5, min_row=data_start_row, max_row=data_end_row)
    series_total = Series(y_total, x_values, title="Total Display Power (W)")
    chart.series.append(series_total)

    y_backlight = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)
    series_bl = Series(y_backlight, x_values, title="Backlight Power (W)")
    chart.series.append(series_bl)

    y_analog = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_end_row)
    series_analog = Series(y_analog, x_values, title="Analog Power (W)")
    chart.series.append(series_analog)

    chart_cell = f"G{table_start_row}"
    ws.add_chart(chart, chart_cell)


def _create_per_brightness_sheets(wb, metrics):
    """Create per-brightness sheets with per-run columns, Average, and Std Dev."""
    summary_table = collections.OrderedDict()
    summary_table_metrics = collections.OrderedDict()

    # Order: Power.* first, then others
    runtype_order_list = []
    for test_name in metrics:
        if "Power." in test_name:
            runtype_order_list.append(test_name)
    for test_name in metrics:
        if test_name not in runtype_order_list:
            runtype_order_list.append(test_name)

    for test_name in runtype_order_list:
        sheet = wb.create_sheet(test_name)
        runtype, _ = test_name.split('.', 1)
        summary_table_metrics.setdefault(runtype, collections.OrderedDict())
        summary_table[test_name] = collections.OrderedDict()
        row = 1
        end_col_num = 0
        conditional_term_set = {'Power', 'MOS', 'Record Time', 'Drain Rate', 'Energy Drained', 'Duration', 'W)'}
        for key in metrics[test_name]:
            data_row = [key] + metrics[test_name][key]
            data_len = len(data_row)
            if data_len > end_col_num:
                end_col_num = data_len

            for i in range(data_len, end_col_num):
                data_row.append("NA")

            start_col_num = 2
            start = get_column_letter(start_col_num) + str(row)
            end = get_column_letter(end_col_num) + str(row)
            avg_col = get_column_letter(end_col_num + 1)
            if row == 1:
                data_row += ['Average', 'Std Dev']
                summary_formula = test_name.split('.', 1)[-1]
            else:
                rng = start + ':' + end
                data_row += ['=IFERROR(AVERAGE(' + rng + '),IF(COUNTIF(' + rng + ',' + start + ')=COUNTA(' + rng + '),' + start + ',"Various"))']
                if any(x in key for x in conditional_term_set):
                    data_row += ['=_xlfn.STDEV.P(' + rng + ')']
                    try:
                        mean = np.mean(metrics[test_name][key])
                        min_val = mean * 0.90
                        max_val = mean * 1.10
                        sheet.conditional_formatting.add(rng,
                            ColorScaleRule(start_type='num', start_value=min_val-0.001, start_color='66CCFF',
                                        mid_type='num', mid_value=mean, mid_color='FFFFFF',
                                        end_type='num', end_value=max_val+0.001, end_color='FF7777'))
                    except Exception:
                        logging.debug("Skipped conditional formatting for %s in %s", key, test_name)
                summary_formula = "='" + test_name + "'!" + avg_col + str(row)
            sheet.append(data_row)
            if "Power" in key or "MOS" in key:
                for col in range(start_col_num, end_col_num + 3):
                    sheet.cell(row, col).number_format = '0.000'
            if "Drain Rate" in key or "Energy Drained" in key or "Record Time" in key:
                for col in range(start_col_num, end_col_num + 3):
                    sheet.cell(row, col).number_format = '0.0'
            summary_table[test_name][key] = summary_formula
            summary_table_metrics[runtype][key] = key
            row += 1

        for col in range(sheet.max_column):
            col_letter = get_column_letter(col + 1)
            if col_letter == 'A':
                sheet.column_dimensions[col_letter].width = 60
            else:
                sheet.column_dimensions[col_letter].width = 15

    return summary_table, summary_table_metrics


def _create_power_summary_sheet(wb, summary_table, summary_table_metrics):
    """Create the Power Summary sheet with formula references to per-brightness sheets."""
    sheet_name = "Power Summary"
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name, 1)
    sheet = wb[sheet_name]

    col = 1
    first_runtype = True
    for compound_name in summary_table:
        runtype, _ = compound_name.split('.', 1)
        if runtype not in summary_table_metrics:
            continue
        row = 1
        if first_runtype:
            for key in summary_table_metrics[runtype]:
                sheet.cell(column=col, row=row).value = summary_table_metrics[runtype][key]
                row += 1
            col += 1
            row = 1
            first_runtype = False

        for key in summary_table_metrics[runtype]:
            if key in summary_table[compound_name]:
                sheet.cell(column=col, row=row).value = summary_table[compound_name][key]
            row += 1
        col += 1

    for c in range(sheet.max_column):
        col_letter = get_column_letter(c + 1)
        if col_letter == 'A':
            sheet.column_dimensions[col_letter].width = 30
        else:
            sheet.column_dimensions[col_letter].width = 15


class BrightnessStudyReport(core.app_scenario.Scenario):
    """
    Standalone scenario that generates a brightness curve study report.
    Produces an XLSX with: Backlight Curve, Power Summary, and per-brightness sheets.
    """
    module = __module__.split('.')[-1]

    Params.setDefault(module, 'result_path', '')
    Params.setDefault(module, 'name', '')
    Params.setDefault(module, 'device_name', '')
    Params.setDefault(module, 'backlight_key', 'DisplayLight Power (W)',
                      desc="Metric name for backlight power in brightness curve report.")
    Params.setDefault(module, 'analog_key', 'DisplayLogic Power (W)',
                      desc="Metric name for analog/panel power in brightness curve report.")

    is_prep = True

    def setUp(self):
        self.result_path = Params.get(self.module, 'result_path')
        self.report_name = Params.get(self.module, 'name')
        self.device_name = Params.get(self.module, 'device_name')
        self.backlight_key = Params.get(self.module, 'backlight_key')
        self.analog_key = Params.get(self.module, 'analog_key')
        return

    def runTest(self):
        if not self.result_path:
            self.result_path = Params.getCalculated("study_result_dir")

        dirpath = os.path.abspath(self.result_path)
        study_name = os.path.basename(dirpath)
        logging.info("Generating Brightness Study Report for: %s", dirpath)

        # Determine report filename
        if self.report_name:
            report_name = self.report_name
        elif self.device_name:
            report_name = study_name + "_" + self.device_name + "_brightness_study_report.xlsx"
        else:
            report_name = study_name + "_brightness_study_report.xlsx"

        # Collect metrics from Brightness-XX folders
        metrics = _collect_metrics(dirpath)
        if not metrics:
            logging.error("No metrics data found in %s", dirpath)
            self.fail("No metrics data found.")
            return

        logging.info("Found %d brightness entries", len(metrics))

        # Create workbook
        wb = openpyxl.Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)

        # Backlight Curve sheet (position 0)
        _create_backlight_curve_sheet(wb, metrics, self.backlight_key, self.analog_key)

        # Per-brightness sheets + collect summary data
        summary_table, summary_table_metrics = _create_per_brightness_sheets(wb, metrics)

        # Power Summary sheet (position 1, after Backlight Curve)
        _create_power_summary_sheet(wb, summary_table, summary_table_metrics)

        # Save report
        output_path = os.path.join(dirpath, report_name)
        wb.save(output_path)
        logging.info("Brightness study report saved to: %s", output_path)

    def tearDown(self):
        return

    def kill(self):
        return 0
