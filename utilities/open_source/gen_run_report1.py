"""
//--------------------------------------------------------------
//
// gen_run_report1
//
// Copyright(c) Microsoft Corporation
// All rights reserved.
//
// MIT License
//
// Permission is hereby granted, free of charge, to any person obtaining
// a copy of this software and associated documentation files(the ""Software""),
// to deal in the Software without restriction, including without limitation the rights
// to use, copy, modify, merge, publish, distribute, sublicense, and / or sell copies
// of the Software, and to permit persons to whom the Software is furnished to do so,
// subject to the following conditions :
//
// The above copyright notice and this permission notice shall be included
// in all copies or substantial portions of the Software.
//
// THE SOFTWARE IS PROVIDED *AS IS*, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
// INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS
// FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.IN NO EVENT SHALL THE AUTHORS
// OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
// WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF
// OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
//
//--------------------------------------------------------------
"""

from openpyxl import load_workbook
import argparse
import glob
import os
import xlwings as xw
import csv

arg_parser = argparse.ArgumentParser(description = "Generate Run Report.")
arg_parser.add_argument('-recurse', '-r', help='Search directories recursively.', action="store_true")
arg_parser.add_argument('-template', '-t', required=True, help='Path to Excel template file.')
arg_parser.add_argument('-report_level', '-l', default="0", help='Level for phase report.')
arg_parser.add_argument('path', nargs='?', default='.\*_DAQ.csv', help='Path to csv DAQ file(s)')
args = arg_parser.parse_args()


def convert_type(str):
    try:
        return int(str)
    except:
        pass

    try:
        return float(str)
    except:
        return (str)


# Function to process ConfigPre/Post files:
def scenario_config(sheet, str):
    print("Writing Config file:  " + str)
    with open(str) as csv_file:
        reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        global count
        for r in reader:
            sheet.cell(row=count, column=4).value = r[0]
            sheet.cell(row=count, column=5).value = (convert_type(r[1]))
            count += 1


# Function to process ConfigPre/Post files:
def scenario_big_config(sheet, str):
    print("Writing Big Config file:  " + str)
    with open(str) as csv_file:
        reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        global count
        for r in reader:
            sheet.cell(row=count, column=1).value = r[0]
            sheet.cell(row=count, column=2).value = (convert_type(r[1]))
            count += 1


# Get directory to process
abspath = os.path.abspath(args.path)
basepath = os.path.basename(abspath) 
if os.path.isdir(abspath):
    dirpath = abspath
    basepath = "*_DAQ.csv"
else:
    dirpath = os.path.dirname(abspath)

# Set the test name to the leaf directory name of the path
testname = dirpath.split("\\")[-1]

print("Processing dir: " + dirpath + ", file spec: " + basepath)
# Process DAQ files
for root, dirs, files in os.walk(dirpath):
    for file in files:
        if glob.fnmatch.fnmatch(file, basepath):
            wb = load_workbook(args.template)

            inputfile = root + os.sep + file

            # Look for Config.csv file, write data if they exist
            sheet = wb["Config"]
            count = 1
            configbig = root + os.sep + "Config.csv"
            if os.path.isfile(configbig):
                scenario_big_config(sheet, configbig)
            else:
                print("Config.csv file not found")

            # Look for ConfigPre and ConfigPost files, write data if they exist
            sheet = wb["Config"]
            count = 1
            configpre = root + os.sep + os.path.basename(root) + "_ConfigPre.csv"
            if os.path.isfile(configpre):
                scenario_config(sheet, configpre)
            else:
                print("ConfigPre file not found")

            configpost = root + os.sep + os.path.basename(root) + "_ConfigPost.csv"
            if os.path.isfile(configpost):
                scenario_config(sheet, configpost)
            else:
                print("ConfigPre file not found")

            inputfile = root + os.sep + file          
       
            # Read in DAQ cvs file
            # Assumes format is:
            #   1st row: channel labels
            #   2nd row: channel data
            daq_data = []
            with open(inputfile) as csvfile:
                reader = csv.reader(csvfile, delimiter=',', quotechar='"')
                for r in reader:
                    daq_data.append(r)

            # Write to Excel file
            try:
                sheet = wb['Raw_Data']
            except:
                sheet = wb['Raw Data']
            row_num = 1
            for row in daq_data:
                col_num = 1
                for val in row:
                    sheet.cell(row=row_num, column=col_num).value = val.strip()
                    col_num += 1
                row_num += 1

            report_name = root + os.sep + testname + "_report.xlsx"
            print("Writing XLSX file:  " + report_name)
            wb.save(report_name)


            # TEMPORARY:  Export Subsystems sheet to power_data.csv
            # Open excel instance
            excel = xw.App(visible = False)

            # Read in Run Report file
            print("Reading Run Report file:  " + report_name)

            wb = xw.books.open(report_name)
            excel = xw.apps.active
            
            sheet = wb.sheets["Subsystems"]
            table = sheet.range('A1').expand().value
            
            excel.quit()

            #Output .csv file of subsystem power data
            csv_name = root + os.sep + testname + "_power_data.csv"
            print("Writing CSV file: " + csv_name)
            with open(csv_name, 'w') as csvfile:
                writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
                for row in table:
                    writer.writerow(row)


    if not args.recurse:
        break

