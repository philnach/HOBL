"""
//--------------------------------------------------------------
//
// gen_run_report2
//
// Copyright(c) Microsoft Corporation
// All rights reserved.
//
// MIT License
//
// Permission is hereby granted, free of charge, to any person obtaining
// a copy of this software and associated documentation files(the ""Software""),
// to deal in the Software without restriction, including without limitation the rights
// to use, copy, modify, merge, publish, distribute, sublicense, and / or sell copies
// of the Software, and to permit persons to whom the Software is furnished to do so,
// subject to the following conditions :
//
// The above copyright notice and this permission notice shall be included
// in all copies or substantial portions of the Software.
//
// THE SOFTWARE IS PROVIDED *AS IS*, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
// INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS
// FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.IN NO EVENT SHALL THE AUTHORS
// OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
// WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF
// OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
//
//--------------------------------------------------------------
"""

from openpyxl import load_workbook
import argparse
import glob
import os
import xlwings as xw
import csv
import pandas as pd

arg_parser = argparse.ArgumentParser(description = "Generate Run Report.")
arg_parser.add_argument('-recurse', '-r', help='Search directories recursively.', action="store_true")
arg_parser.add_argument('-report_level', '-l', default="0", help='Level for phase report.')
arg_parser.add_argument('-template', '-t', required=True, help='Path to Excel template file.')
arg_parser.add_argument('path', nargs='?', default='.\log.csv', help='Path to csv DAQ file(s)')
args = arg_parser.parse_args()


def convert_type(str):
    try:
        return int(str)
    except:
        pass

    try:
        return float(str)
    except:
        return (str)


# Function to process ConfigPre/Post files:
def scenario_config(sheet, str):
    print("Writing Config file:  " + str)
    with open(str) as csv_file:
        reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        global count
        for r in reader:
            sheet.cell(row=count, column=4).value = r[0]
            sheet.cell(row=count, column=5).value = (convert_type(r[1]))
            count += 1


# Get directory to process
abspath = os.path.abspath(args.path)
basepath = os.path.basename(abspath) 
if os.path.isdir(abspath):
    dirpath = abspath
    basepath = "log.csv"
else:
    dirpath = os.path.dirname(abspath)

# Find latest config_check folder
config_name = ""
found_cfg = False
for parent_dir in [dirpath, os.path.dirname(dirpath), os.path.dirname(os.path.dirname(dirpath))]:
    print("Looking for config_check_??? in: " + parent_dir)
    for dir in os.listdir(parent_dir):
        if glob.fnmatch.fnmatch(dir, "config_check_???"):
            config_name = parent_dir + os.sep + dir + os.sep + "Config.csv"
            print("Found Config.csv at : " + config_name)
            found_cfg = True
    if found_cfg:
        break
    # break - Must comment/delete to find LATEST config.csv

# Reads Big Config data into table - only once!
test_table = []
if config_name != "":

    # print("Creating big Config table")
    with open(config_name) as csv_file:
        reader = csv.reader(csv_file, delimiter=',', quotechar='"')
        for row in reader:
            test_table.append(row)
else:
    print("ERROR:  Could not find a config_check_??? run.")

# Set the test name to the leaf directory name of the path
testname = dirpath.split("\\")[-1]

print("Processing dir: " + dirpath + ", file spec: " + basepath)
# Process DAQ files
for root, dirs, files in os.walk(dirpath):
    for file in files:
        if glob.fnmatch.fnmatch(file, basepath):
            wb = load_workbook(args.template)
            # Write big Config data
            if config_name != "":
                sheet = wb["Config"]
                table_count = 1
                for row in test_table:
                    sheet.cell(row=table_count, column=1).value = row[0]
                    sheet.cell(row=table_count, column=2).value = (convert_type(row[1]))
                    table_count += 1

            inputfile = root + os.sep + file

            # Look for ConfigPre and ConfigPost files, write data if they exist
            sheet = wb["Config"]
            count = 1
            configpre = root + os.sep + os.path.basename(root) + "_ConfigPre.csv"
            if os.path.isfile(configpre):
                scenario_config(sheet, configpre)
            else:
                print("ConfigPre file not found")

            configpost = root + os.sep + os.path.basename(root) + "_ConfigPost.csv"
            if os.path.isfile(configpost):
                scenario_config(sheet, configpost)
            else:
                print("ConfigPre file not found")

            raw_file = root + os.sep + file        

            # Convert raw log.csv file to a summary file that just has the average of all the channels
            summary_file = root + os.sep + "summary_DAQ.csv"
            df=pd.read_csv(raw_file, sep=',', header=5)
            dfA = df.mean().to_frame().T
            dfA.to_csv(summary_file, sep=',', index=False, float_format='%0.6f')

            # Read in DAQ summary csv file
            # Assumes format is:
            #   1st row: channel labels
            #   2nd row: channel data
            daq_data = []
            with open(summary_file) as csvfile:
                reader = csv.reader(csvfile, delimiter=',', quotechar='"')
                for r in reader:
                    daq_data.append(r)

            # Write to Excel file
            wb = load_workbook(args.template)
            try:
                sheet =  wb['Raw_Data']
            except:
                sheet = wb['Raw Data']
            row_num = 1
            for row in daq_data:
                col_num = 1
                for val in row:
                    try:
                        sheet.cell(row = row_num, column = col_num).value = float(val.strip())
                    except:
                        sheet.cell(row = row_num, column = col_num).value = val.strip()
                    sheet.cell(row = row_num, column = col_num).number_format = '0.000000'
                    col_num += 1
                row_num += 1

            report_name = root + os.sep + testname + "_report.xlsx"
            print("Writing XLSX file:  " + report_name)
            wb.save(report_name)


            # TEMPORARY:  Export Subsystems sheet to power_data.csv
            # Open excel instance
            excel = xw.App(visible = False)

            # Read in Run Report file
            print("Reading Run Report file:  " + report_name)

            wb = xw.books.open(report_name)
            excel = xw.apps.active
            
            sheet = wb.sheets["Subsystems"]
            table = sheet.range('A1').expand().value
            
            excel.quit()

            #Output .csv file of subsystem power data
            csv_name = root + os.sep + testname + "_power_data.csv"
            print("Writing CSV file: " + csv_name)
            with open(csv_name, 'w') as csvfile:
                writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
                for row in table:
                    writer.writerow(row)


    if not args.recurse:
        break

