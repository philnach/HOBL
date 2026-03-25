"""
//--------------------------------------------------------------
//
// Generate HOBL Trend
//
// Copyright(c) Microsoft Corporation
// All rights reserved.
//
// MIT License
//
// Permission is hereby granted, free of charge, to any person obtaining
// a copy of this software and associated documentation files(the ""Software""),
// to deal in the Software without restriction, including without limitation the rights
// to use, copy, modify, merge, publish, distribute, sublicense, and / or sell copies
// of the Software, and to permit persons to whom the Software is furnished to do so,
// subject to the following conditions :
//
// The above copyright notice and this permission notice shall be included
// in all copies or substantial portions of the Software.
//
// THE SOFTWARE IS PROVIDED *AS IS*, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
// INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS
// FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.IN NO EVENT SHALL THE AUTHORS
// OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
// WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF
// OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
//
//--------------------------------------------------------------
"""


import sys
import argparse
import glob
import os
# import xlwings as xw
import csv
import collections
import time
import pandas as pd
import collections
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def process(path, unified, refresh):
    # Get directory to work with
    dirpath = os.path.abspath(path)
    basepath = "*study_report.xlsx"
    csv_folder = "study_report_csv"
    subdir_filter = ["PostLaunch\\23H2_22631", "PreLaunch\\24H2_26100", "PreLaunch\\24H2_26090"]
    subdir_filter = ["."]
    output_list = []
    rollup_start_time = datetime.now()

    if unified == "1":
        full_csv_folder = dirpath + os.sep + csv_folder
        if not os.path.exists(full_csv_folder):
            os.makedirs(full_csv_folder)

        # Read in config table, if exists, to get report modification timestamps

        # config
        csv_name = full_csv_folder + os.sep + "config.csv"
        csv_mtime = os.path.getmtime(csv_name) - 300  # subtract 5 min to accomodate reports that are new during the time it took to do the last rollup
        if os.path.exists(csv_name) and refresh == '0':
            attributes = set({"Report", "Timestamp"})
            config_df = pd.read_csv(csv_name, index_col="Report")
            print("Found existing config table")
        else:
            attributes = set({"Report", "Timestamp"})
            config_df = pd.DataFrame()
            print("Creating new Config table")
        config_file_list = config_df.index.tolist()
        stage2_config_df = pd.DataFrame()

        # runs
        csv_name = full_csv_folder + os.sep + "runs.csv"
        if os.path.exists(csv_name) and refresh == '0':
            metrics = set({"Report", "Run Type", "Run Number"})
            runs_df = pd.read_csv(csv_name, index_col="Report")
            print("Found existing Runs table")
        else:
            metrics = set({"Report", "Run Type", "Run Number"})
            runs_df = pd.DataFrame()
            print("Creating new Runs table")
        stage2_runs_df = pd.DataFrame()

        # # sd
        # csv_name = full_csv_folder + os.sep + "sd.csv"
        # if os.path.exists(csv_name) and refresh == '0':
        #     metrics = set({"Report", "Run Type", "Run Number"})
        #     sd_df = pd.read_csv(csv_name, index_col="Report")
        #     print("Found existing SD table")
        # else:
        #     metrics = set({"Report", "Run Type", "Run Number"})
        #     sd_df = pd.DataFrame()
        #     print("Creating new SD table")

        # # hobl
        # csv_name = full_csv_folder + os.sep + "hobl.csv"
        # if os.path.exists(csv_name) and refresh == '0':
        #     study_table = []
        #     with open(csv_name) as csvfile:
        #         reader = csv.reader(csvfile, delimiter=',', quotechar='"')
        #         for r in reader:
        #             study_table.append(r)
        #     print("Found existing HOBL table")
        # else:
        #     study_table = [["Report", "Study", "Active On", "Active On Target", "Telemetry", "Telemetry Target", "HOBL", "HOBL Target"]]
        #     print("Creating new HOBL table")

        # # scorecard
        # csv_name = full_csv_folder + os.sep + "scorecard.csv"
        # if os.path.exists(csv_name) and refresh == '0':
        #     scorecard_table = []
        #     with open(csv_name) as csvfile:
        #         reader = csv.reader(csvfile, delimiter=',', quotechar='"')
        #         for r in reader:
        #             scorecard_table.append(r)
        #     print("Found existing Scorecard table")
        # else:
        #     scorecard_table = [["Report", "Scenario Name", "Scenario", "Sub-system", "Target", "Measured", "Delta", "Scenario Index", "Subsystem Index"]]
        #     print("Creating new Scorecard table")
        

    # Process study_report files
    # excel = xw.App(visible = False)

    for subdir in subdir_filter:
        if subdir == ".":
            sub_path = dirpath
        else:
            sub_path = dirpath + os.sep + subdir
        print("Processing dir: " + sub_path + ", file spec: " + basepath)
        report_found = False
        for root, dirs, files in os.walk(sub_path, topdown=True):
            if os.path.exists(root + os.sep + "ignore_trend.txt"):
                print("Ignoring " + root)
                continue
            if refresh == '0':
                dirs[:] = [d for d in dirs if os.path.getmtime(os.path.join(root, d)) >= csv_mtime]
            try:
                dirs.remove("Training")
            except:
                pass
            try:
                dirs.remove("Prep")
            except:
                pass

            print(root, dirs, files)
            for file in files:
                if glob.fnmatch.fnmatch(file, basepath):
                    if file[0] == "~":
                        continue
                    report_found = True
                    # Once we've found a report, no need to look in sub-directories
                    dirs[:] = []
                    inputfile = root + os.sep + file
                    print("Reading Study Report file:  " + inputfile)
                    study_name = inputfile.split('\\')[-2]

                    mtime = os.path.getmtime(inputfile)
                    time_str = time.strftime(" %Y-%m-%d %H:%M:%S", time.localtime(mtime))
                    print("Mod time: ", time_str)

                    # if inputfile in config_df.index.tolist():
                    if inputfile in config_file_list:
                        csv_time_str = config_df.at[inputfile, "Timestamp"]
                        print("CSV time: ", csv_time_str)
                        csv_secs = time.mktime(time.strptime(csv_time_str, " %Y-%m-%d %H:%M:%S"))
                        if (csv_secs + 1) > mtime and refresh == '0':
                            # Report not modified so leave alone
                            print("Study: " + study_name + " Not modified.")
                            continue
                        else:
                            if unified == "1":
                                # Report modified, delete this report from all tables and reprocess
                                if inputfile in config_df.index:
                                    config_df.drop(inputfile, inplace=True)
                                if inputfile in runs_df.index:
                                    runs_df.drop(inputfile, inplace=True)
                                # if inputfile in sd_df.index:
                                #     sd_df.drop(inputfile, inplace=True)
                                # study_table = [row for row in study_table if row[0] != inputfile]
                                # scorecard_table = [row for row in scorecard_table if row[0] != inputfile]


                    # Read in Study Report file
                    try:
                        # wb = excel.books.open(inputfile)
                        wb = openpyxl.load_workbook(inputfile, read_only=True)

                    except Exception as e:
                        print("ERROR - could not open " + inputfile)
                        print(f"EXCEPTION: {str(e)}")
                        continue

                    # sheet_names = [s.name for s in wb.sheets]
                    sheet_names = wb.sheetnames

                    # if unified == "0":
                    #     full_csv_folder = root + os.sep + csv_folder
                    #     if not os.path.exists(full_csv_folder):
                    #         os.makedirs(full_csv_folder)

                    ##
                    # Config
                    ##

                    if "Config" in sheet_names:
                    # if False:
                        # if unified == "0":
                        #     attributes = set({"Report"})
                        #     config_df = pd.DataFrame()

                        sheet = wb["Config"]
                        # keys = sheet.range('A2').expand('down').value
                        keys = []
                        for row in sheet.iter_rows(min_row=1, max_col=1):
                            keys.append(row[0].value)
                        y = len(keys) + 1
                        keys = ["Report", "Timestamp"] + keys
                        # attributes.update((keys))

                        vals = []
                        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2):
                            vals.append(row[0].value)

                        # vals = sheet['B1:B' + str(y)]
                        string_vals = [inputfile, time_str] + [v if v != None else "" for v in vals]
                        
                        d = collections.OrderedDict(zip(keys, string_vals))
                        new_config_df = pd.DataFrame([d])
                        new_config_df = new_config_df.set_index('Report')
                        # print new_config_df
                        stage2_config_df = pd.concat([stage2_config_df, new_config_df], sort=False)
                        # print config_df

                        # if unified == "0":
                        #     csv_name = root + os.sep + csv_folder + os.sep + "config.csv"
                        #     print("Writing CSV file: " + csv_name)
                        #     with open(csv_name, 'wb') as csvfile:
                        #         csv_name = root + os.sep + csv_folder + os.sep + "config.csv"
                        #         print("Writing CSV file: " + csv_name)
                        #         config_df.to_csv(csv_name, sep=',')

                        if "Config" not in output_list:            
                            output_list.append("Config")
                    else:
                        print("ERROR: Can't rollup a report without a Config sheet.")
                        print(inputfile)
                        print("Quitting.")
                        exit(1)

                    ##
                    # HOBL
                    ##

                    # if "HOBL" in sheet_names:
                    # # if False:
                    #     sheet = wb.sheets["HOBL"]

                    #     study_report_version = sheet.range('A1').value
                    #     if study_report_version == None or study_report_version == "":
                    #         # Search for keys in column B, values in F
                    #         for y in range(20,40):
                    #             if sheet.range('B' + str(y)).value == "Screen On Battery Life Prediction":
                    #                 break
                    #         print("Found Active On at row " + str(y))
                    #         active_val = sheet.range('F'+ str(y)).value
                    #         hobl_val = sheet.range('F'+ str(y+2)).value
                    #         lvp_val = sheet.range('F'+ str(y+3)).value
                    #         active_target = sheet.range('L52').value
                    #         lvp_target = sheet.range('N52').value
                    #         hobl_target = sheet.range('P52').value
                    #     elif int(study_report_version) == 2:
                    #         # Search for keys in column C, values in G
                    #         for y in range(20,40):
                    #             if sheet.range('C' + str(y)).value == "Screen On Battery Life Prediction (h)":
                    #                 break
                    #         print("Found Active On at row " + str(y))
                    #         active_val = sheet.range('G'+ str(y)).value
                    #         hobl_val = sheet.range('G'+ str(y+4)).value
                    #         lvp_val = sheet.range('G'+ str(y+5)).value
                    #         active_target = sheet.range('L64').value
                    #         lvp_target = sheet.range('N64').value
                    #         hobl_target = sheet.range('P64').value
                    #     elif int(study_report_version) == 3:
                    #         # Search for keys in column C, values in G
                    #         for y in range(20,40):
                    #             if sheet.range('C' + str(y)).value == "Screen On Battery Life Prediction (h)":
                    #                 break
                    #         print("Found Active On at row " + str(y))
                    #         active_val = sheet.range('G'+ str(y)).value
                    #         hobl_val = sheet.range('G'+ str(y+5)).value
                    #         lvp_val = sheet.range('G'+ str(y+6)).value
                    #         active_target = sheet.range('L64').value
                    #         lvp_target = sheet.range('N64').value
                    #         hobl_target = sheet.range('P64').value
                    #     elif int(study_report_version) == 4:
                    #         # Search for keys in column C, values in G
                    #         for y in range(20,40):
                    #             if sheet.range('C' + str(y)).value == "Screen On Battery Life Prediction (h)":
                    #                 break
                    #         print("Found Active On at row " + str(y))
                    #         active_val = sheet.range('G'+ str(y)).value
                    #         hobl_val = sheet.range('G'+ str(y+4)).value
                    #         lvp_val = sheet.range('G'+ str(y+7)).value
                    #         active_target = sheet.range('L64').value
                    #         lvp_target = sheet.range('N64').value
                    #         hobl_target = sheet.range('P64').value
                    #     elif int(study_report_version) >= 5:
                    #         # Search for keys in column C, values in G
                    #         for y in range(20,40):
                    #             if sheet.range('C' + str(y)).value == "HOBL Active On (h)":
                    #                 break
                    #         print("Found Active On at row " + str(y))
                    #         active_val = sheet.range('G'+ str(y)).value
                    #         hobl_val = sheet.range('G'+ str(y+5)).value
                    #         abl_val = sheet.range('G'+ str(y+6)).value
                    #         perf_val = sheet.range('G'+ str(y+7)).value
                    #         active_target = sheet.range('L65').value
                    #         lvp_target = 0
                    #         hobl_target = sheet.range('N65').value

                    #     try:
                    #         float(active_val)
                    #     except:
                    #         active_val = 0
                    #     try:
                    #         float(hobl_val)
                    #     except:
                    #         hobl_val = 0
                    #     try:
                    #         float(active_target)
                    #     except:
                    #         active_target = 0
                    #     try:
                    #         float(lvp_target)
                    #     except:
                    #         lvp_target = 0
                    #     try:
                    #         float(hobl_target)
                    #     except:
                    #         hobl_target = 0
                    #     try:
                    #         float(lvp_val)
                    #     except:
                    #         lvp_val = 0
                    #     try:
                    #         float(abl_val)
                    #     except:
                    #         abl_val = 0
                    #     try:
                    #         float(perf_val)
                    #     except:
                    #         perf_val = 0


                    #     if unified == "0":
                    #         study_table = []
                    #         study_table.append(["Report", "Study", "Active On", "Active On Target", "HOBL", "HOBL Target", "ABL Lab", "ABL Lab Perf"])
                    #     study_table.append([
                    #         inputfile, 
                    #         study_name, 
                    #         "%.1f" % active_val, 
                    #         "%.1f" % float(active_target), 
                    #         "%.1f" % hobl_val, 
                    #         "%.1f" % float(hobl_target),
                    #         "%.1f" % float(abl_val), 
                    #         "%.1f" % float(perf_val)
                    #     ])

                    #     if unified == "0":
                    #         csv_name = root + os.sep + csv_folder + os.sep + "hobl.csv"
                    #         print("Writing CSV file: " + csv_name)
                    #         with open(csv_name, 'wb') as csvfile:
                    #             writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    #             for row in study_table:
                    #                 writer.writerow(row)
                    #     if "HOBL" not in output_list:            
                    #         output_list.append("HOBL")

                    # ##
                    # # Scorecard
                    # ##

                    # if "Scorecard" in sheet_names:
                    # # if False:
                    #     if unified == "0":
                    #         scorecard_table = [["Report", "Scenario Name", "Scenario", "Sub-system", "Target", "Measured", "Delta", "Scenario Index", "Subsystem Index"]]
                    #     sheet = wb.sheets["Scorecard"]
                    #     # Use row 3 to count columns because upper rows have expanded columns
                    #     types = sheet.range('B3').expand('right').value
                    #     columns = len(types)
                    #     end_col = columns + 1
                    #     data_start_row = 4
                    #     raw_scenarios = sheet.range((1,2), (1, end_col)).value
                    #     scenarios = sheet.range((2,2), (2, end_col)).value
                    #     subsystems = sheet.range('A4').expand('down').value
                    #     for row in range(len(subsystems)):
                    #         data_row = data_start_row + row
                    #         data = sheet.range((data_row, 2), (data_row, end_col)).value
                    #         for col in range(columns):
                    #             valtype = types[col]
                    #             if valtype == "Measured":
                    #                 scenario_index = col - 1
                    #                 scenario = scenarios[col - 1]
                    #                 raw_scenario = raw_scenarios[col - 1]
                    #                 measured_value = data[col]
                    #                 target_value = data[col - 1]
                    #                 try:
                    #                     delta = "%.2f" % (measured_value / target_value)
                    #                 except:
                    #                     delta = None
                    #                 subsystem = subsystems[row]
                    #                 line = [inputfile, raw_scenario, scenario, subsystem, target_value, measured_value, delta, scenario_index, row]
                    #                 scorecard_table.append(line)

                    #     if unified == "0":
                    #         csv_name = root + os.sep + csv_folder + os.sep + "scorecard.csv"
                    #         print("Writing CSV file: " + csv_name)
                    #         with open(csv_name, 'wb') as csvfile:
                    #             writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    #             for row in scorecard_table:
                    #                 writer.writerow(row)

                    #     if "Scorecard" not in output_list:            
                    #         output_list.append("Scorecard")


                    ##
                    # Runs
                    ##

                    # if unified == "0":
                    #     metrics = set({"Report", "Run Type", "Run Number"})
                    #     runs_df = pd.DataFrame()
                    stage_runs_df = pd.DataFrame([d])
                    stage_runs_df = stage_runs_df.set_index('Report')

                    for sheet_name in sheet_names:
                        if "." in sheet_name:
                            run_type, scenario = sheet_name.rsplit(".", 1)
                            if run_type != "Power":
                                continue
                            sheet = wb[sheet_name]
                            # keys = sheet.range('A2').expand('down').value
                            # keys = sheet['A']
                            keys = []
                            for row in sheet.iter_rows(min_row=2, max_col=1):
                                keys.append(row[0].value)
                            if keys == None:
                                continue
                            y = len(keys) + 1
                            keys = ["Report", "Run Type", "Run Number"] + keys
                            # metrics.update((keys))
                            # print metrics
                            col = 2
                            while True:
                                # print("Processing sheet: " + sheet_name)
                                # run_number = sheet.range(1, col).value
                                run_number = sheet.cell(row=1, column=col).value
                                run_number = str(run_number)
                                # print("Processing run_number: " + run_number)
                                if run_number == "Average" or not run_number.isdigit():
                                    break
                                
                                # vals = sheet[(2,col), (y,col)]
                                # vals = sheet[get_column_letter(col)]
                                # if len(vals) > 1:
                                #     vals.pop(0)
                                vals = []
                                for row in sheet.iter_rows(min_row=2, max_row=y, min_col=col, max_col=col):
                                    vals.append(row[0].value)

                                # no_enter_vals = [v if '\n' not in v else v.split('\n')[0] for v in vals]
                                no_enter_vals = []
                                for v in vals:
                                    try:
                                        v1 = v.splitlines()[0]
                                    except:
                                        v1 = v
                                    try:
                                        v2 = v1.split(',')[0]
                                    except:
                                        v2 = v1
                                    no_enter_vals.append(v2)
                                # no_enter_vals = [v if "\n" not in v else v.splitlines()[0] for v in vals]
                                # no_comma_vals = [v if ',' not in v else v.split(',')[0] for v in no_enter_vals]
                                string_vals = [v if v != None and v != "NA" else "" for v in no_enter_vals]
                                string_vals = [inputfile, run_type, run_number] + string_vals
                                d = collections.OrderedDict(zip(keys, string_vals))
                                new_runs_df = pd.DataFrame([d])
                                new_runs_df = new_runs_df.set_index('Report')
                                # start_time = datetime.now()
                                stage_runs_df = pd.concat([stage_runs_df, new_runs_df], sort=False)
                                # print(f"Time delta: {datetime.now() - start_time}")
                                col += 1

                            if "Runs" not in output_list:
                                print ("Adding to Runs")            
                                output_list.append("Runs")
                    # start_time2 = datetime.now()
                    stage2_runs_df = pd.concat([stage2_runs_df, stage_runs_df], sort=False)
                    # print(f"Time delta2: {datetime.now() - start_time2}")


                    # if unified == "0":
                    # # if False:
                    #     csv_name = root + os.sep + csv_folder + os.sep + "runs.csv"
                    #     print("Writing CSV file: " + csv_name)
                    #     runs_df.transpose().to_csv(csv_name, sep=',')


                    ##
                    # SD
                    ##

                    # if unified == "0":
                    #     metrics = set({"Report", "Run Type", "Run Number"})
                    #     sd_df = pd.DataFrame()
                    # for sheet_name in sheet_names:
                    #     if "." in sheet_name:
                    #         run_type, scenario = sheet_name.rsplit(".", 1)
                    #         if run_type != "SD":
                    #             continue
                    #         sheet = wb.sheets[sheet_name]
                    #         keys = sheet.range('A2').expand('down').value
                    #         y = len(keys) + 1
                    #         keys = ["Report", "Run Type", "Run Number"] + keys
                    #         metrics.update((keys))
                    #         # print metrics
                    #         col = 2
                    #         while True:
                    #             # print "Processing sheet: " + sheet_name
                    #             run_number = sheet.range(1, col).value
                    #             run_number = str(run_number)
                    #             # print "Processing run_number: " + run_number
                    #             if run_number == "Average" or not run_number.isdigit():
                    #                 break
                                
                    #             vals = sheet.range((2,col), (y,col)).value
                    #             # if "NA" in vals:
                    #             #     col += 1
                    #             #     continue
                    #             string_vals = [v if v != None else "" for v in vals]
                    #             string_vals = [inputfile, run_type, run_number] + string_vals
                    #             d = collections.OrderedDict(zip(keys, string_vals))
                    #             new_sd_df = pd.DataFrame([d])
                    #             new_sd_df = new_sd_df.set_index('Report')
                    #             sd_df = pd.concat([sd_df, new_sd_df], sort=False)
                    #             col += 1

                    #         if "SD" not in output_list:            
                    #             print ("Adding to SD")            
                    #             output_list.append("SD")

                    # if unified == "0":
                    # # if False:
                    #     csv_name = root + os.sep + csv_folder + os.sep + "sd.csv"
                    #     print("Writing CSV file: " + csv_name)
                    #     sd_df.transpose().to_csv(csv_name, sep=',')


                    # wb.close()

    if unified == "1":
        # # Write hobl
        # if "HOBL" in output_list:
        #     csv_name = dirpath + os.sep + csv_folder + os.sep + "hobl.csv"
        #     print("Writing CSV file: " + csv_name)
        #     with open(csv_name, 'w') as csvfile:
        #         writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        #         for row in study_table:
        #             writer.writerow(row)
        # # Write scorecard
        # if "Scorecard" in output_list:
        #     csv_name = dirpath + os.sep + csv_folder + os.sep + "scorecard.csv"
        #     print("Writing CSV file: " + csv_name)
        #     with open(csv_name, 'w') as csvfile:
        #         writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        #         for row in scorecard_table:
        #             writer.writerow(row)
        # Write config
        if "Config" in output_list:
            config_df = pd.concat([config_df, stage2_config_df], sort=False)
            csv_name = dirpath + os.sep + csv_folder + os.sep + "config.csv"
            print("Writing CSV file: " + csv_name)
            config_df.to_csv(csv_name, sep=',')
        # Write runs
        if "Runs" in output_list:
            runs_df = pd.concat([runs_df, stage2_runs_df], sort=False)
            csv_name = dirpath + os.sep + csv_folder + os.sep + "runs.csv"
            print("Writing CSV file: " + csv_name)
            runs_df.to_csv(csv_name, sep=',')
        # # Write sd
        # if "SD" in output_list:
        #     csv_name = dirpath + os.sep + csv_folder + os.sep + "sd.csv"
        #     print("Writing CSV file: " + csv_name)
        #     sd_df.to_csv(csv_name, sep=',')

    print("Quitting.")
    # excel.quit()
    print(f"Rollup took: {datetime.now() - rollup_start_time}s")

def main():
    # Parse command line arguments
    arg_parser = argparse.ArgumentParser(description = "Generate Run Report.")
    arg_parser.add_argument('path', nargs='?', default='.', help='Path to studies.')
    arg_parser.add_argument('-unified', '-u', nargs='?', default='1', help='Separate files or one big singele one?')
    arg_parser.add_argument('-refresh', '-r', nargs='?', default='0', help='Reprocess all study reports')
    args = arg_parser.parse_args()

    process(args.path, args.unified, args.refresh)

if __name__ == "__main__":
    main()

