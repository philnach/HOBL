"""
//--------------------------------------------------------------
//
// gen_study_report_2017
//
// Copyright(c) Microsoft Corporation
// All rights reserved.
//
// MIT License
//
// Permission is hereby granted, free of charge, to any person obtaining
// a copy of this software and associated documentation files(the ""Software""),
// to deal in the Software without restriction, including without limitation the rights
// to use, copy, modify, merge, publish, distribute, sublicense, and / or sell copies
// of the Software, and to permit persons to whom the Software is furnished to do so,
// subject to the following conditions :
//
// The above copyright notice and this permission notice shall be included
// in all copies or substantial portions of the Software.
//
// THE SOFTWARE IS PROVIDED *AS IS*, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
// INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS
// FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.IN NO EVENT SHALL THE AUTHORS
// OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
// WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF
// OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
//
//--------------------------------------------------------------
"""

import openpyxl
from openpyxl.chart import (
    Reference,
    Series,
    LineChart,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.chart.text import RichText
from openpyxl.drawing.colors import ColorChoice
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import xlwings as xw

import sys
import argparse
import glob
import os
import shutil
import collections
import csv
import pandas as pd
import numpy as np
import utilities.open_source.gen_hobl_trend as gen_hobl_trend
import time
import subprocess


def main():
    arg_parser = argparse.ArgumentParser(description = "Generate Study Report.")
    arg_parser.add_argument('-template', '-t', default='docs\\hobl_study_report_template.xlsx', help='Path to Excel template file.')
    arg_parser.add_argument('-trend', '-tr', help='Path to hobl trend .csv file for populating trend plot.')
    arg_parser.add_argument('-goals', '-g', help='Path to scenario goals .csv file for populating goals in table.')
    arg_parser.add_argument('-name', '-n', default='study_report.xlsx', help='Name of the report file.')
    arg_parser.add_argument('-adders', '-d', help='Path to scenario adders .csv file for populating touch and keyboard adders in table.')
    arg_parser.add_argument('-product', '-p', help='Name of product (copied to config sheet)')
    arg_parser.add_argument('-study_type', '-s', help='Type of study (copied to config sheet)')
    arg_parser.add_argument('-device_name', '-dn', help='Name of the device (copied to config sheet)')
    arg_parser.add_argument('-comments', '-c', help='Comments about this study (copied to config sheet)')
    arg_parser.add_argument('-html_only', '-ho', nargs='?', default='0', help='Create an HTML report, but not the .xlsx report.')
    arg_parser.add_argument('-active_target', '-a', nargs='?', default='0', help='Active Target in hours.')
    arg_parser.add_argument('-hobl_target', '-o', nargs='?', default='0', help='HOBL Target in hours.')
    arg_parser.add_argument('-battery_capacity', '-b', nargs='?', default='0', help='Typical battery capacity.')
    arg_parser.add_argument('-battery_min_capacity', '-bm', nargs='?', help='Minimum battery capacity.')
    arg_parser.add_argument('-battery_reserve', '-r', nargs='?', default='0', help='Fraction of battery capacity not able to be consumed.')
    arg_parser.add_argument('-battery_derating', '-bd', nargs='?', help='Fraction of battery capacity not able to be consumed.')
    arg_parser.add_argument('-hibernate_timeout', '-ht', nargs='?', default='6', help='Hibernate Timeout in hours.')
    arg_parser.add_argument('-hibernate_budget_target', '-hbt', nargs='?', default='0', help='Hibernate Budget Target in percent.')
    arg_parser.add_argument('-enable_phase_report', '-pr', nargs='?', default='1', help='Enable or disable phase reports.')
    arg_parser.add_argument('path', nargs='?', default='.\\*metrics.csv', help='Path to metrics.csv files')
    args = arg_parser.parse_args()

    # Get directory to process
    abspath = os.path.abspath(args.path)
    basepath = os.path.basename(abspath)
    if os.path.isdir(abspath):
        dirpath = abspath
        basepath = "*metrics.csv"
    else:
        dirpath = os.path.dirname(abspath)
    study_name = dirpath.split('\\')[-1]
    print (u"Study name: " + study_name)
    if args.device_name != "" and args.device_name != None :
        study_name = study_name + "_" + args.device_name
    if args.name == "study_report.xlsx":
        args.name = study_name + "_study_report.xlsx"

    metrics = collections.OrderedDict()


    def host_call(command, cwd = "."):
        print("Calling: " + command)
        p = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell = True, cwd = cwd)
        out, err = p.communicate()
        for line in out.split(b'\n'):
            print(line.decode().rstrip())
        for line in err.split(b'\n'):
            print(line.decode().rstrip())
        return(out.decode())


    def convert_type(str):
        try:
            return int(str)
        except:
            pass
        
        try:
            return float(str)
        except:
            return(str)


    # Read Run metrics .csv files
    for root, dirs, files in os.walk(dirpath):
        parent = os.path.abspath(root + os.sep + "..")
        parent_parent = os.path.abspath(parent + os.sep + "..")
        # if not os.path.exists(root + os.sep + ".PASS") or os.path.exists(root + os.sep + ".FAIL"):
        #     continue

        if not os.path.exists(root + os.sep + ".PASS") and not os.path.exists(parent + os.sep + ".PASS") and not os.path.exists(parent_parent + os.sep + ".PASS"):
            continue

        if args.enable_phase_report is None:
            args.enable_phase_report = "1"

        if os.path.exists(parent_parent + os.sep + ".PASS") and (args.enable_phase_report).lower() == "0":
            continue 

        for file in files:
            for file in (root, file):
                path = os.path.join(root,file)
            #for file in files:
                if glob.fnmatch.fnmatch(file, basepath):
                    if "config_check" in os.path.dirname(path) or "training" in os.path.dirname(path) or "misc" in os.path.dirname(path) or "prep" in os.path.dirname(path) or "fail" in os.path.dirname(path):
                        continue
                    inputfile = root + os.sep + file
                    # parent = os.path.abspath(root + os.sep + "..")
                    if parent == dirpath:
                        # Not in any subdirectory, default runtype to "Power"
                        runtype = "Power"
                    else:
                        # Use grandparent of file to determine runtype
                        # runtype = os.path.basename(parent)
                        if os.path.exists(root + os.sep + ".PASS"):
                            runtype = os.path.basename(parent)
                        else:
                            runtype = os.path.basename(os.path.dirname(parent_parent))
                    run_name = file[:-12]
                    test_name = run_name[:-4]
                    compound_name = runtype + "." + test_name
                    # print run_name, test_name
                    if compound_name not in metrics:
                        metrics[compound_name] = collections.OrderedDict()
                    # Read in metrics file
                    print (u"Reading metrics file:  " + inputfile)
                    if 'Metric' not in metrics[compound_name]:
                        metrics[compound_name]['Metric'] = []
                    metrics[compound_name]['Metric'].append(run_name[-3:]) # Just the run number
                    num_runs = len(metrics[compound_name]['Metric'])
                    with open(inputfile) as csvfile:
                        reader = csv.reader(csvfile, delimiter=',', quotechar='"')
                        for r in reader:
                            key = r[0]
                            val = convert_type(r[1])
                            if key not in metrics[compound_name]:
                                metrics[compound_name][key] = []
                            num_this_metric = len(metrics[compound_name][key])
                            for i in range(num_this_metric, num_runs-1):
                                metrics[compound_name][key].append("NA")
                            metrics[compound_name][key].append(val)




    # Write to Excel Study Report file
    print (u"Opening template file: " + args.template)
    wb = openpyxl.load_workbook(args.template)
    summary_table = collections.OrderedDict()
    summary_table_metrics = collections.OrderedDict()

    # Create summary sheets so they appear before the scenario sheets
    summary_sheet_list = []
    for compound_name in metrics:
        runtype, testname = compound_name.split('.', 1)
        sheet_name = runtype + " Summary"
        if sheet_name in summary_sheet_list:
            continue
        summary_sheet_list.append(sheet_name)
        if sheet_name in wb.sheetnames:
            # Clear cells of any existing sheets
            for row in wb[sheet_name]['A1:Z100']:
                for cell in row:
                    cell.value = None
        else:
            wb.create_sheet(sheet_name)

    # # Populate Config sheet
    if "Config" not in wb.sheetnames:
        wb.create_sheet("Config")
    sheet = wb["Config"]
    # Find latest config folder
    config_name = ""
    config_root = ""
    for root, dirs, files in os.walk(dirpath):
        for dir in dirs:
            if glob.fnmatch.fnmatch(dir, "config_check_???"):
                config_name = dir
                config_root = root
    if config_name != "":
        config_name = config_root + os.sep + config_name + os.sep + "config.csv"
        print("CONFIG")
        print (config_name)
        with open(config_name) as csv_file:
            reader = csv.reader(csv_file, delimiter=',', quotechar='"')
            for r in reader:
                sheet.append([r[0], convert_type(r[1])])
        # Adjust column widths
        for col in range(sheet.max_column):
            col_letter = get_column_letter(col + 1)
            sheet.column_dimensions[col_letter].width = 40
    else:
        print("Config file not found. Leaving it blank.")
    #    print (" ERROR - Could not find a config_check_??? run.")
    #    sys.exit(-1)

    # Print out the power data first in the study report
    runtype_order_list = []
    for test_name in metrics:
        if "Power." in test_name:
            runtype_order_list.append(test_name)
    for test_name in metrics:
        if test_name not in runtype_order_list:
            runtype_order_list.append(test_name)

    # summary_table_metrics = collections.OrderedDict()
    for test_name in runtype_order_list:
        runtype, shortname = test_name.split('.', 1)
        if (runtype not in summary_table_metrics):
            summary_table_metrics[runtype] = collections.OrderedDict()

    for test_name in runtype_order_list:
        wb.create_sheet(test_name)
        sheet = wb[test_name]
        runtype, shortname = test_name.split('.', 1)
        summary_table[test_name] = collections.OrderedDict()
        row = 1
        end_col_num = 0
        for key in metrics[test_name]:
            # Add formula for average
            data_row = [key] + metrics[test_name][key]
            data_len = len(data_row)
            if data_len > end_col_num:
                end_col_num = data_len
            
            for i in range(data_len, end_col_num):
                data_row.append("NA")

            start_col_num = 2
            col = openpyxl.utils.get_column_letter(start_col_num)
            start = col + str(row)
            col = openpyxl.utils.get_column_letter(end_col_num)
            end = col + str(row)
            avg_col = openpyxl.utils.get_column_letter(end_col_num + 1)
            if row == 1:
                data_row += ['Average', 'Std Dev']
                summary_formula = test_name.split('.', 1)[-1]
            else:
                rng = start + ':' + end
                data_row += ['=IFERROR(AVERAGE(' + rng + '),IF(COUNTIF(' + rng + ',' + start + ')=COUNTA(' + rng + '),' + start + ',"Various"))']
                conditional_term_set = set(['Power', 'MOS', 'Record Time', 'Drain Rate', 'Energy Drained', 'Duration', 'W)'])
                # Check if any of the above conditional formatting desired terms is in the current key, and apply conditional formatting if any come back True
                if any([x in key for x in conditional_term_set]):
                    data_row += ['=_xlfn.STDEV.P(' + rng + ')']
                    # Add a three-color scale
                    # print test_name, key, metrics[test_name][key]
                    # if "" in metrics[test_name][key] or "NA" in metrics[test_name][key]:
                    #     continue
                    try:
                        mean = np.mean(metrics[test_name][key])
                        min_val = mean * 0.90
                        max_val = mean * 1.10
                        sheet.conditional_formatting.add(rng,
                            ColorScaleRule(start_type='num', start_value=min_val-0.001, start_color='66CCFF',
                                        mid_type='num', mid_value=mean, mid_color='FFFFFF',
                                        end_type='num', end_value=max_val+0.001, end_color='FF7777')
                                        )
                    except:
                        pass
                summary_formula = "='" + test_name + "'!" + avg_col + str(row)
            sheet.append(data_row)
            if "Power" in key or "MOS" in key:
                for col in range(start_col_num, end_col_num + 3):
                    sheet.cell(row, col).number_format = '0.000'
            if "Drain Rate" in key or "Energy Drained" in key or "Record Time" in key:
                for col in range(start_col_num, end_col_num + 3):
                    sheet.cell(row, col).number_format = '0.0'
            summary_table[test_name][key] = summary_formula
            summary_table_metrics[runtype][key] = key
            row += 1
        # Adjust column widths
        for col in range(sheet.max_column):
            col_letter = get_column_letter(col + 1)
            if col_letter == 'A':
                sheet.column_dimensions[col_letter].width = 60
            else:
                sheet.column_dimensions[col_letter].width = 15

    # Populate Summary sheets
    for summary_sheet_name in summary_sheet_list:
        col = 1
        # Populate each column
        for compound_name in summary_table:
            runtype, testname = compound_name.split('.', 1)
            sheet_runtype = summary_sheet_name.replace(" Summary", "")
            if runtype == sheet_runtype:
                sheet = wb[summary_sheet_name]
                row = 1
                if col == 1:
                    # Populate column 1 with metric keys
                    for key in summary_table_metrics[runtype]:
                        sheet.cell(column = col, row = row).value = summary_table_metrics[runtype][key]
                        row += 1
                    col += 1
                    row = 1
                # Populate value columns by looping through keys in column 1
                for key in summary_table_metrics[runtype]:
                    if key in summary_table[compound_name]:
                        sheet.cell(column = col, row = row).value = summary_table[compound_name][key]
                    row += 1
                col += 1
        # Adjust column widths
        for col in range(sheet.max_column):
            col_letter = get_column_letter(col + 1)
            if col_letter == 'A':
                sheet.column_dimensions[col_letter].width = 30
            else:
                sheet.column_dimensions[col_letter].width = 15

    if args.goals:
        # Read goals csv into dictionary
        # goals = pd.read_csv(args.goals, index_col=0, sep='\t').to_dict(orient= 'dict')
        if "Scorecard" in wb:
            goals = pd.read_csv(args.goals, index_col=0, sep=',')
            print(goals)
            sheet = wb["Scorecard"]
            # Write subsystem names to Scorecard
            subsystem_row = 4
            for subsystem_name in goals.index: 
                sheet.cell(row = subsystem_row, column = 1).value = subsystem_name 
                subsystem_row += 1 
            # Process Scorecard subsystems
            total_power_found = False 
            for subsystem_cell in sheet['A']: 
                row = subsystem_cell.row 
                if row == 1 or row == 2:
                    continue
                if str(subsystem_cell.value) == "None":
                    break
                if total_power_found == True and "Record Time" not in str(subsystem_cell.value):
                    sheet.delete_rows(row, 1)
                if "Total Power" in str(subsystem_cell.value):
                    total_power_found = True
            
            # Fixes VLOOKUP formula of the last row 
            start = '$' 
            end = ',' 
            for cell in sheet[(row-1)]: 
                if "=IFERROR(VLOOKUP" in str(cell.value): 
                    cell_value = str(cell.value) 
                    new_value = cell_value.replace(str((cell_value.split(start))[1].split(end)[0]), "A"+str(row-1)) 
                    cell.value = new_value
                    cell.data_type = 'f'

            # Formats Total Power and Record Time rows with borders
            border_side = Side(border_style="medium", color="FFFFFF")
            border_bottom = Side(border_style="medium")
            for cell in sheet[row-2:row-2]: 
                cell.border = Border(top=Side(border_style="double"), bottom=border_bottom, left=border_side, right=border_side)
            for cell in sheet[row-1:row-1]:
                cell.border = Border(bottom=border_bottom, left=border_side, right=border_side) 

            # Writes goal numbers into Scorecard
            for scenarios_cell in sheet['1']: 
                scenario = scenarios_cell.value 
                col = scenarios_cell.column   
                for subsystem_cell in sheet['A']:
                    subsystem = subsystem_cell.value
                    row = subsystem_cell.row 
                    if scenario in goals and subsystem in goals[scenario]: 
                        sheet.cell(row = row, column = col).value = goals[scenario][subsystem]
        else:
            print("No Scorecard tab found, skipping goals.")

    if args.adders:
        # Read adders csv into dictionary
        # CSV file should contain scenario names in top row, and adder metrics in column 0
        adders = pd.read_csv(args.adders, index_col=0, sep=',')
        print (adders)

        study_report_version = wb["HOBL"].cell(row = 1, column = 1).value #checks the report version number

        adder_type_row = '17'
        if study_report_version >= 9:
            adder_type_row = '18'
        elif study_report_version >= 7:
            adder_type_row = '21'

        sheet = wb["Scenarios"]
        for adder_type_cell in sheet[adder_type_row]:
            adder_type = adder_type_cell.value
            if adder_type is None:
                continue
            col = adder_type_cell.col_idx
            for scenario_cell in sheet['D']:
                scenario = scenario_cell.value
                row = scenario_cell.row
                valid = 0
                if scenario in adders and adder_type in adders[scenario]:
                    if pd.isnull(adders[scenario][adder_type]):
                        continue
                    sheet.cell(row = row, column = col).value = adders[scenario][adder_type]

    if "HOBL" in wb:
        if args.battery_capacity:
            sheet = wb["HOBL"]
            sheet.cell(row = 20, column = 7).value = float(args.battery_capacity)
            sheet.cell(row = 20, column = 7).number_format = '0.0'

        if args.battery_min_capacity:
            sheet = wb["HOBL"]
            sheet.cell(row = 19, column = 7).value = float(args.battery_min_capacity)
            sheet.cell(row = 19, column = 7).number_format = '0.0'

        if args.battery_derating:
            sheet = wb["HOBL"]
            sheet.cell(row = 21, column = 8).value = float(args.battery_derating)
            sheet.cell(row = 21, column = 8).number_format = '0%'

        if args.battery_reserve:
            sheet = wb["HOBL"]
            sheet.cell(row = 22, column = 8).value = float(args.battery_reserve)
            sheet.cell(row = 22, column = 8).number_format = '0%'

        # if args.hibernate_timeout:
        #     sheet = wb["HOBL"]
        #     sheet.cell(row = 32, column = 7).value = float(args.hibernate_timeout)
        #     sheet.cell(row = 32, column = 7).number_format = '0'

        if args.hibernate_budget_target:
            sheet = wb["HOBL"]
            sheet.cell(row = 33, column = 7).value = float(args.hibernate_budget_target)
            sheet.cell(row = 33, column = 7).number_format = '0%'

    if args.product:
        try:
            sheet = wb["Config"]
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, max_row=60):
                cell = row[0]
                if cell.value == "Product":
                    sheet.cell(row = cell.row, column = 2).value = args.product
                    break
        except:
            print("Could not write Product to Config sheet.")

    if args.study_type:
        try:
            sheet = wb["Config"]
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, max_row=60):
                cell = row[0]
                if cell.value == "Study Type":
                    sheet.cell(row = cell.row, column = 2).value = args.study_type
                    break
        except:
            print("Could not write Study Type to Config sheet.")

    if args.device_name:
        try:
            sheet = wb["Config"]
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, max_row=60):
                cell = row[0]
                if cell.value == "Device Name":
                    sheet.cell(row = cell.row, column = 2).value = args.device_name
                    break
        except:
            print("Could not write Device Name to Config sheet.")

    if args.comments:
        try:
            sheet = wb["Config"]
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, max_row=60):
                cell = row[0]
                if cell.value == "Comments":
                    sheet.cell(row = cell.row, column = 2).value = args.comments
                    break
        except:
            print("Could not write Comments to Config sheet.")


    if args.trend:
        # Only process trend of HOBL sheet exists
        if "HOBL" in wb:
            # Update trend CSV with any new data
            gen_hobl_trend.process(args.trend, args.active_target, args.hobl_target)

            sheet = wb["HOBL"]
            # Get trend data
            trend_data = []
            with open(args.trend + os.sep + "trend_data.csv") as csvfile:
                dialect = csv.Sniffer().sniff(csvfile.read(1024))
                csvfile.seek(0)
                reader = csv.reader(csvfile, dialect)
                for r in reader:
                    if r[0] == study_name:
                        continue
                    trend_data.append(list(r))

            max_rows = 13 # Number of rows in speadsheet template table
            num_rows = len(trend_data) - 1 # -1 for header
            trend_start_row = 1
            padding_rows = 0
            if num_rows > max_rows:
                trend_start_row = num_rows - max_rows + 2
            elif num_rows < max_rows:
                padding_rows = max_rows - num_rows
            
            # Header
            study_report_version = sheet.cell(row = 1, column = 1).value #checks the report version number
            if study_report_version <= 2:
                row = 51
                col = 10
            else:
                row = 52
                col = 10

            for val in trend_data[0]:
                if val == "Timestamp":
                    continue
                sheet.cell(row = row, column = col).value = val
                col += 1
            
            # Trend data
            row += 1
            # print num_rows, max_rows, trend_start_row
            for trend_row in range(trend_start_row, trend_start_row + min([num_rows, max_rows]) - 1):
                # print "TREND ROW: ", trend_row
                r = trend_data[trend_row]
                if r[0] == study_name:
                    continue
                col = 10
                for val in r:
                    if col == 10:
                        sheet.cell(row = row, column = col).value = val
                        sheet.cell(row = row, column = col).alignment = openpyxl.styles.alignment.Alignment(horizontal='right')
                    else:
                        if (val == ""):
                            sheet.cell(row = row, column = col).value = val
                            sheet.cell(row = row, column = col).number_format = '0.0'
                        else:
                            sheet.cell(row = row, column = col).value = float(val)
                            sheet.cell(row = row, column = col).number_format = '0.0'
                    col += 1
                    if col > 16:
                        break
                row += 1
                
            if study_report_version == None or study_report_version == "":
                sheet.cell(row = row, column = 11).value = "=F30" # Active On
                sheet.cell(row = row, column = 13).value = "=F33" # LVP
                sheet.cell(row = row, column = 15).value = "=F32" # HOBL
            elif study_report_version == 2:
                sheet.cell(row = row, column = 11).value = "=G30" # Active On
                sheet.cell(row = row, column = 13).value = "=G35" # LVP
                sheet.cell(row = row, column = 15).value = "=G34" # HOBL
            elif study_report_version == 3:
                sheet.cell(row = row, column = 11).value = "=G30" # Active On
                sheet.cell(row = row, column = 13).value = "=G36" # LVP
                sheet.cell(row = row, column = 15).value = "=G35" # HOBL
            elif study_report_version == 4:
                sheet.cell(row = row, column = 11).value = "=G30" # Active On
                sheet.cell(row = row, column = 13).value = "=G37" # LVP
                sheet.cell(row = row, column = 15).value = "=G34" # HOBL
            elif study_report_version == 5:
                sheet.cell(row = row, column = 11).value = "=G30" # Active On
                sheet.cell(row = row, column = 13).value = "=G35" # HOBL
            elif study_report_version >= 6:
                sheet.cell(row = row, column = 11).value = "=G37" # ABL Active On
                sheet.cell(row = row, column = 13).value = "=G35" # HOBL
            
            # For last row of table, reference this sheet's calculated Active On, LVP, and HOBL scores
            sheet.cell(row = row, column = 10).value = study_name
            sheet.cell(row = row, column = 10).alignment = openpyxl.styles.alignment.Alignment(horizontal='right')
            # sheet.cell(row = row, column = 11).value = "=G30"  # Active On
            sheet.cell(row = row, column = 11).number_format = '0.0'
            sheet.cell(row = row, column = 12).value = float(args.active_target)
            sheet.cell(row = row, column = 12).number_format = '0.0'
            # # sheet.cell(row = row, column = 13).value = float(args.telemetry_value)
            # sheet.cell(row = row, column = 13).number_format = '0.0'
            # sheet.cell(row = row, column = 14).value = float(args.telemetry_target)
            # sheet.cell(row = row, column = 14).number_format = '0.0'
            # sheet.cell(row = row, column = 15).value = "=G34"  # HOBL
            sheet.cell(row = row, column = 13).number_format = '0.0'
            sheet.cell(row = row, column = 14).value = float(args.hobl_target)
            sheet.cell(row = row, column = 14).number_format = '0.0'

            for trend_row in range(padding_rows):
                sheet.cell(row = row, column = 12).value = float(args.active_target)
                sheet.cell(row = row, column = 12).number_format = '0.0'
                # sheet.cell(row = row, column = 14).value = float(args.telemetry_target)
                # sheet.cell(row = row, column = 14).number_format = '0.0'
                sheet.cell(row = row, column = 14).value = float(args.hobl_target)
                sheet.cell(row = row, column = 14).number_format = '0.0'
                row += 1

            # # Plot chart
            # chart_data = Reference(sheet, min_col=11, min_row=51, max_col=16, max_row=row)
            # cat_data = Reference(sheet, min_col=10, min_row=52, max_col=10, max_row=row)
            # chart = LineChart()
            # chart.height = 15
            # chart.width = 30
            # chart.add_data(data=chart_data, titles_from_data=True)
            # chart.set_categories(cat_data)
            # chart.dataLabels = DataLabelList()
            # chart.dataLabels.showVal = True
            # chart.dataLabels.dLblPos = 'ctr'
            # cp = CharacterProperties(solidFill=ColorChoice(prstClr='white'), sz=700) # White, 7pt text
            # chart.dataLabels.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            # chart.y_axis.title = 'Hours of Battery Life'
            # chart.x_axis.title = 'OS Build'
            # cp = CharacterProperties(sz=700) # 7pt text
            # chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
            # chart.y_axis.delete=True
            # # Format data lines
            # color_list = ["ED7D31", "ED7D31", "A5A5A5", "A5A5A5", "4472C4", "4472C4"]
            # for s in [0, 2, 4]:
            #     line = chart.series[s]
            #     line.marker.symbol = "circle"
            #     line.marker.size = 17
            #     line.graphicalProperties.line.solidFill = color_list[s]
            #     line.marker.graphicalProperties.solidFill = color_list[s] # Marker filling
            #     line.marker.graphicalProperties.line.solidFill = color_list[s] # Marker outline
            #     line.graphicalProperties.line.width = 31750 # width in EMUs
            # # Format target lines
            # for s in [1, 3, 5]:
            #     line = chart.series[s]
            #     line.dLbls = DataLabelList()
            #     line.dLbls.showVal = False
            #     line.graphicalProperties.line.solidFill = color_list[s]
            #     line.graphicalProperties.line.dashStyle = "dash"
            #     line.graphicalProperties.line.width = 31750 # width in EMUs

            # sheet.add_chart(chart, "A51")

    # Save report
    excel_file = dirpath + os.sep + args.name
    html_name = excel_file.replace(".xlsx", ".htm")
    html_files = excel_file.replace(".xlsx", "_files")
    if args.html_only == "1":
        excel_file = dirpath + os.sep + "temp_report.xlsx"
    wb.save(excel_file)
    print (u"Writing " + excel_file)

    # Convert to HTML
    print("Converting Study Report to HTML:  " + excel_file)

    # host_call("powershell.exe utilities\\excel_to_html.ps1 " + excel_file + " " + html_name)
    # sys.exit(0)


    # Open excel instance
    # excel_opened = False
    # excel = xw.App(visible = True)
    # excel_opened = True
    # print ("Excel opened")
    # wb = excel.books.open(excel_file)
    # sys.exit(0)

    excel_opened = False
    try:
        excel = xw.App(visible = False)
        excel_opened = True
    except:
        print("Unable to open Excel to convert report to HTML.  Is it installed?")

    # Read in Run Report file
    if excel_opened:
        xwb = excel.books.open(excel_file)
        # excel = xw.apps.active
        try:
            print("Removing HTML study report files:  " + html_files)
            shutil.rmtree(html_files)
        except:
            print ("Could not delete HTML study report files.")
        try:
            print("Removing HTML study report:  " + html_name)
            os.remove(html_name)
        except:
            print ("Could not delete HTML study report.")
        try:
            # A little delay to prevent "can't save" error.
            time.sleep(2)
            print("Writing HTML to:  " + html_name)
            xwb.api.SaveAs(html_name, xw.constants.FileFormat.xlHtml, xw.constants.SaveConflictResolution.xlLocalSessionChanges)
        except:
            print ("ERROR: unable to save html.")
        time.sleep(2)
        print("Quitting Excel")
        excel.quit()

    if args.html_only == "1":
        try:
            time.sleep(2)
            os.remove(excel_file)
        except:
            print ("Could not remove temp excel file.")

    sys.exit(0)

if __name__ == "__main__":
    main()
